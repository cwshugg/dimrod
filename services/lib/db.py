# This module implements wrapper code around an SQLite3 database, to make it
# easier to work with.
#
#   Connor Shugg

import os
import sys
import sqlite3
import openpyxl

# Enable import from the parent directory
pdir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
if pdir not in sys.path:
    sys.path.append(pdir)

from lib.config import Config, ConfigField

class DatabaseConfig(Config):
    """Represent a database configuration."""
    def __init__(self):
        super().__init__()
        self.fields = [
            ConfigField("path",             [str],      required=True),
            # `check_same_thread` — forwarded to `sqlite3.connect`. Defaults to
            # `True` (sqlite3's own default). Set to `False` to allow the
            # cached connection to be used from threads other than the creator.
            # When doing so the CALLER is responsible for serializing access
            # (e.g. with a lock), exactly as sqlite3 requires.
            ConfigField("check_same_thread", [bool],    required=False, default=True),
            # `timeout` — forwarded to `sqlite3.connect` (seconds to wait on a
            # locked database before raising). When `None` sqlite3's default
            # (5.0s) is used, i.e. the connection is created without passing
            # the argument at all — identical to the historical behavior.
            ConfigField("timeout",           [float],   required=False, default=None),
            # `busy_timeout_ms` — when set, a `PRAGMA busy_timeout` is issued
            # on every freshly-created connection. When `None` no pragma is
            # issued (historical behavior).
            ConfigField("busy_timeout_ms",   [int],     required=False, default=None),
        ]

class Database:
    """Represent a database interface."""
    def __init__(self, config: DatabaseConfig):
        """Initializes a new database object with the provided config.

        Connection behavior (thread-safety / timeouts) is driven entirely by
        the `DatabaseConfig` fields `check_same_thread`, `timeout` and
        `busy_timeout_ms`. When a config is built from just `{"path": ...}`
        those fields resolve to their defaults (True/None/None), reproducing the
        historical `sqlite3.connect(self.config.path)` behavior exactly.
        """
        self.config = config
        self.conn = None

    def _new_connection(self):
        """Creates a brand new sqlite3 connection honoring the config's
        thread-safety/timeout options. Existing callers whose config sets none
        of these options get exactly `sqlite3.connect(self.config.path)` as
        before.
        """
        # Read the connection options off the config. getattr guards against a
        # config object that predates these fields.
        check_same_thread = getattr(self.config, "check_same_thread", True)
        timeout = getattr(self.config, "timeout", None)
        busy_timeout_ms = getattr(self.config, "busy_timeout_ms", None)

        kwargs = {}
        # Only deviate from sqlite3's own defaults when explicitly asked to, so
        # existing callers observe byte-for-byte identical connection behavior.
        if check_same_thread is not True:
            kwargs["check_same_thread"] = check_same_thread
        if timeout is not None:
            kwargs["timeout"] = timeout
        conn = sqlite3.connect(self.config.path, **kwargs)
        if busy_timeout_ms is not None:
            conn.execute("PRAGMA busy_timeout = %d" % int(busy_timeout_ms))
        return conn

    def get_connection(self, reset=False):
        """Retrieves the current connection cached in the object, or creates a new
        one if it doesn't exist.
        """
        # If we don't have a connection, create one.
        if self.conn is None:
            self.conn = self._new_connection()

        # If we already have a connection but a reset was requested, close the
        # existing connection and create a new one.
        if self.conn is not None and reset:
            self.conn.close()
            self.conn = self._new_connection()

        return self.conn

    def close_connection(self):
        """Closes the object's cached connection."""
        if self.conn is not None:
            self.conn.close()
            self.conn = None

    def delete(self):
        """Deletes the databsae file, completely wiping all data and removing it from
        the filesystem.
        """
        self.close_connection()
        if os.path.exists(self.config.path):
            os.remove(self.config.path)

    def execute(self, query: str, do_commit=False, params: tuple = None):
        """Executes the provided query and returns the result of
        `connection.execute()`.

        If `do_commit` is set to `True`, the transaction will be committed after
        executing the query.

        When `params` is provided (a tuple/list), the query is executed with
        bound parameters (`?` placeholders), which is the SQL-safe way to pass
        untrusted values. When `params` is `None` the query is executed
        exactly as before with no binding (historical behavior).
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        if params is not None:
            result = cursor.execute(query, params)
        else:
            result = cursor.execute(query)

        # If requested, commit the transaction.
        if do_commit:
            conn.commit()
        return result

    def table_exists(self, table: str) -> bool:
        """Determines if a table exists in the database."""
        conn = self.get_connection()
        cur = conn.cursor()
        result = cur.execute("SELECT 1 FROM sqlite_master WHERE type == 'table' AND name == '%s';" % table)
        table_exists = result.fetchone() is not None
        return table_exists

    def get_all_table_names(self):
        """Returns a list of all tables present in the database."""
        conn = self.get_connection()
        cur = conn.cursor()
        result = cur.execute("SELECT name FROM sqlite_master WHERE type == 'table';")
        return [row[0] for row in result]

    def get_table_column_names(self, table: str):
        """Returns a list of all column names in the provided table."""
        conn = self.get_connection()
        cur = conn.cursor()
        result = cur.execute("PRAGMA table_info(%s);" % table)
        return [row[1] for row in result]

    def search(self, table: str, condition: str, order_by: str = None, desc: bool = False, limit: int = None, params: tuple = None):
        """Performs a search of the database and returns tuples in a list.

        Optionally supports ORDER BY and LIMIT clauses. When `params` is
        provided, `condition` should use `?` placeholders and the values
        are bound safely via parameterized execution.
        """
        # If the table doesn't exist, return an empty list
        if not self.table_exists(table):
            return []

        # Build a SELECT command:
        cmd = "SELECT * FROM %s" % table
        if condition is not None and len(condition) > 0:
            cmd += " WHERE %s" % condition
        if order_by is not None:
            cmd += " ORDER BY %s" % order_by
            if desc:
                cmd += " DESC"
        if limit is not None:
            cmd += " LIMIT %d" % limit

        # Connect, query, and return
        conn = self.get_connection()
        cur = conn.cursor()
        if params is not None:
            result = cur.execute(cmd, params)
        else:
            result = cur.execute(cmd)
        return result

    def search_order_by(self,
                        table: str,
                        order_by_column: str,
                        desc: bool = False,
                        limit: int = None):
        """Executes a search using `ORDER BY` to retrieve entries without needing a
        specific condition to identify them.
        """
        # If the table doesn't exist, return an empty list
        if not self.table_exists(table):
            return []

        # Build a SELECT command
        cmd = "SELECT * FROM %s ORDER BY %s" % (table, order_by_column)
        if desc:
            cmd += " DESC"
        if limit is not None:
            cmd += " LIMIT %d" % limit

        # Connect, query, and return
        conn = self.get_connection()
        cur = conn.cursor()
        result = cur.execute(cmd)
        return result

    def table_to_csv(self, table: str, condition: str):
        """Queries the database and returns a table's values, filtered using
        `condition` as a CSV string.
        """
        result = self.search(table, condition)
        return "\n".join([",".join(row) for row in result])

    def export_to_excel(self, path: str, table_names: list[str] = None):
        """Exports all tables present (or only the ones specified in `table_names`)
        in the database to an Excel (spreadsheet) file at the provided path.
        """
        wb = openpyxl.Workbook()

        # If no table names were provided, export all tables.
        if table_names is None:
            table_names = self.get_all_table_names()

        # For each table...
        for table_name in table_names:
            # Create a new sheet with the table's name as the title:
            ws = wb.create_sheet(title=table_name)

            # Retrieve ALL entries in the table:
            result = self.search(table_name, None)

            # Get the column names for the table, and write them to the first
            # row of the sheet.
            column_names = self.get_table_column_names(table_name)
            row_index = 1
            for i, column_name in enumerate(column_names):
                ws.cell(row=row_index, column=i+1).value = column_name
            row_index += 1

            # For each row, write the values to the corresponding cells in the
            # sheet, in the rows directly underneath the column names.
            for i, row in enumerate(result):
                for j, value in enumerate(row):
                    ws.cell(row=i+row_index, column=j+1).value = value

        # Save the workbook to the provided path.
        wb.save(path)

    def insert_or_replace(self, table: str, values, do_commit: bool = True,
                          columns=None):
        """Inserts a row into the specified table, or replaces it if a row with the
        same primary key already exists.

        Two calling conventions are supported:

        1. `values` is a **string** (historical behavior): it is treated as a
           SQL-formatted values string (e.g. the output of
           `Uniserdes.to_sqlite3_str()`) and interpolated directly into the
           statement. This path is UNCHANGED for existing callers.

        2. `values` is a **tuple/list** (new, SQL-safe path): the values are
           bound via `?` placeholders instead of being interpolated, so
           arbitrary/untrusted content (JSON blobs, user strings, ...) is
           persisted safely. An optional `columns` iterable may be supplied to
           emit an explicit column list, e.g. `INSERT OR REPLACE INTO t
           (a, b) VALUES (?, ?)`.
        """
        conn = self.get_connection()
        cursor = conn.cursor()

        # New parameterized path: bind a sequence of values safely.
        if isinstance(values, (tuple, list)):
            placeholders = ", ".join(["?"] * len(values))
            col_clause = ""
            if columns is not None:
                col_clause = " (%s)" % ", ".join(columns)
            cursor.execute(
                "INSERT OR REPLACE INTO %s%s VALUES (%s)" %
                (table, col_clause, placeholders),
                tuple(values)
            )
        else:
            # Historical string-interpolation path (unchanged).
            cursor.execute("INSERT OR REPLACE INTO %s VALUES %s" % (table, values))

        if do_commit:
            conn.commit()

    def delete_by_column(self, table: str, column: str, value,
                         do_commit: bool = True):
        """Deletes all rows from `table` whose `column` equals `value`.

        This is a small, general, SQL-safe convenience wrapper over
        `delete_where`: the value is always bound via a `?` placeholder so
        callers never build a SQL literal themselves. `column` is a trusted
        identifier (a schema column name), while `value` may be arbitrary /
        untrusted content and is passed as a bound parameter.
        """
        self.delete_where(table, "%s = ?" % column, (value,),
                          do_commit=do_commit)

    def delete_where(self, table: str, condition: str, params: tuple = None,
                     do_commit: bool = True):
        """Deletes rows from `table` matching `condition`.

        When `params` is provided, `condition` should use `?` placeholders
        and the values are bound safely via parameterized execution — this is
        the SQL-safe way to delete by an untrusted key. When `params` is
        `None` the condition is used verbatim.
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        cmd = "DELETE FROM %s" % table
        if condition is not None and len(condition) > 0:
            cmd += " WHERE %s" % condition
        if params is not None:
            cursor.execute(cmd, params)
        else:
            cursor.execute(cmd)
        if do_commit:
            conn.commit()

